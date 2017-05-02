#rateValidation.py - Checks if a rate input sheet is valid.

#Modules
import openpyxl, datetime, re, ibm_db

#Global Constants
validStates = ['AL', 'AK', 'AZ', 'AR', 'CA', 'CO', 'CT', 'DE', 'DC', 'FL', 'GA', 'HI', 'ID', 'IL', 'IN', 'IA', 'KS',
               'KY', 'LA', 'ME', 'MD', 'MA', 'MI', 'MN', 'MS', 'MO', 'MT', 'NE', 'NV', 'NH', 'NJ', 'NM', 'NY', 'NC',
               'ND', 'OH', 'OK', 'OR', 'PA', 'PR', 'RI', 'SC', 'SD', 'TN', 'TX', 'UT', 'VT', 'VA', 'WA', 'WV', 'WI',
               'WY', 'OT', '***']

conn = ibm_db.connect('DATABASE=DSNDP30;HOSTNAME=DSNDP30;PORT=2646;PROTOCOL=TCPIP;UID=E0015EY;PWD=acTrpj83;', '', '')
sqlT024X1 = 'SELECT DISTINCT DCLR_INT_RT_TBL FROM E0015DB.T024X WHERE COMPANY_CODE = \'MLF\''
sqlT024X2 = 'SELECT DISTINCT GUAR_INT_RT_TBL FROM E0015DB.T024X WHERE COMPANY_CODE = \'MLF\''
sqlTAD3F = 'SELECT DISTINCT DECL_TBL_NUMBER FROM E0015DB.TAD3F WHERE COMPANY_CODE = \'MLF\''
sqlT0C8X1 = 'SELECT DISTINCT CAP_RATE_TBL FROM E0015DB.T0C8X WHERE COMPANY_CODE = \'MLF\''
sqlT0C8X2 = 'SELECT DISTINCT TRIGGER_RATE_TBL FROM E0015DB.T0C8X WHERE COMPANY_CODE = \'MLF\''
sqlT0C8X3 = 'SELECT DISTINCT GUAR_CAP_RATE_TBL FROM E0015DB.T0C8X WHERE COMPANY_CODE = \'MLF\''
sqlT0C8X4 = 'SELECT DISTINCT GUAR_TRIG_RATE_TBL FROM E0015DB.T0C8X WHERE COMPANY_CODE = \'MLF\''
sqlT0C8X5 = 'SELECT DISTINCT GUAR_SPREAD_TBL FROM E0015DB.T0C8X WHERE COMPANY_CODE = \'MLF\''
field24X1 = 'DCLR_INT_RT_TBL'
field24X2 = 'GUAR_INT_RT_TBL'
fieldD3F = 'DECL_TBL_NUMBER'
fieldT0C8X1 = 'CAP_RATE_TBL'
fieldT0C8X2 = 'TRIGGER_RATE_TBL'
fieldT0C8X3 = 'GUAR_CAP_RATE_TBL'
fieldT0C8X4 = 'GUAR_TRIG_RATE_TBL'
fieldT0C8X5 = 'GUAR_SPREAD_TBL'

#Global Variables
subsetT024X1 = []
subsetT024X2 = []
subsetTAD3F = []
subsetT0C8X1 = []
subsetT0C8X2 = []
subsetT0C8X3 = []
subsetT0C8X4 = []
subsetT0C8X5 = []

#Functions
def populateSubsets(inputConn, inputSQL, inputSubset, inputField):
    stmt = ibm_db.exec_immediate(inputConn, inputSQL)
    dictionary = ibm_db.fetch_assoc(stmt)
    while dictionary != False:
        inputSubset.append(str(dictionary[inputField]).rstrip())
        dictionary = ibm_db.fetch_assoc(stmt)

def validateTableSubset(inputCellValue, inputRow, inputSubset, inputSubset2 = None, inputSubset3 = None, inputSubset4 = None):
    subsetError = []
    if len(str(inputCellValue)) > 16:
        subsetError.append('The TABLE_SUBSET record is too long. Row ' + str(inputRow) + '.')
    else:
        if str(inputCellValue) not in inputSubset:
            if inputSubset2 != None and str(inputCellValue) not in inputSubset2:
                if inputSubset3 != None and str(inputCellValue) not in inputSubset3:
                    if inputSubset4 != None and str(inputCellValue) not in inputSubset4:
                        subsetError.append('The TABLE_SUBSET record is invalid. Row ' + str(inputRow) + '.')
                else:
                    subsetError.append('The TABLE_SUBSET record is invalid. Row ' + str(inputRow) + '.')
            else:
                subsetError.append('The TABLE_SUBSET record is invalid. Row ' + str(inputRow) + '.')
    return subsetError

def printErrorMessages(errorList):
    errorListOutput = ''
    
    if errorList == []: #End of validation, print results to terminal
            return('Validation is complete. No errors were found. \n')
    else:
        return('Validation is complete. The following errors were found: \n' + '\n'.join(errorList) + '\n')
    
def validateControlRecord(inputSheet):
    controlRecordErrors = []
    if inputSheet['A2'].value != inputSheet.title: #Check TABLE_NAME
        controlRecordErrors.append('The TABLE_NAME record is invalid. Cell A2.')

    #if inputSheet['B2'].number_format != 'mm/dd/yyyy': #DATE format check
        #controlRecordErrors.append('The DATE record is not correctly formatted. Cell B2.')

    try: #Validate cycle date
        cycleDate = str(inputSheet['B2'].value)
        datetime.datetime.strptime(cycleDate[0:10], '%Y-%m-%d')
    except:
        controlRecordErrors.append('The DATE record is an invalid date. Cell B2.') 

    if inputSheet.max_row - 3 != inputSheet['C2'].value: #Check RECORD_COUNT
        controlRecordErrors.append('The RECORD_COUNT record is incorrect. Cell C2.')

    return controlRecordErrors

def validateTable(inputSheet):
    if inputSheet.max_row == 1 and inputSheet.max_column == 1: #Blank spreadsheet check
        return None
    elif inputSheet.title == 'T025X': #Return the table that needs to be validated
        return 'T025X'
    elif inputSheet.title == 'T026X':
        return 'T026X'
    elif inputSheet.title == 'TU130':
        return 'TU130'
    else:
        return False

def validateCompanyCode(sheetObject, inputRow, inputColumn):
    companyCodeError = []
    if sheetObject.cell(row = inputRow, column = inputColumn).value != 'MLF':
        companyCodeError.append('An incorrect COMPANY_CODE has been entered. Row ' + str(inputRow) + '.')
    return companyCodeError

def validateProductPrefix(sheetObject, inputRow, inputColumn):
    prodPrefixError = []
    if len(str(sheetObject.cell(row = inputRow, column = inputColumn).value)) > 1:
        prodPrefixError.append('The PRODUCT_PREFIX record is too long. Row ' + str(inputRow) + '.')
    return prodPrefixError

def validateIssueState(sheetObject, inputRow, inputColumn):
    issueStateError = []
    if sheetObject.cell(row = inputRow, column = inputColumn).value not in validStates: #State check
        issueStateError.append('The ISSUE_STATE record is invalid. Row ' + str(inputRow) + '.')
    return issueStateError

def validateData(inputSheet):
    dataErrors = []
    allAddRows = []
    rowData = []
    
    if inputSheet.max_row == 3: #Check to see if data is present, should be outside of for loop to reduce number of checks in loop
        dataErrors.append('No data is present in the spreadsheet.')
    elif inputSheet.title == 'T025X' and inputSheet.max_column != 9:
        dataErrors.append('An incorrect number of columns are present in the spreadsheet. Further validation cannot occur.')
    elif inputSheet.title == 'T026X' and inputSheet.max_column != 9:
        dataErrors.append('An incorrect number of columns are present in the spreadsheet. Further validation cannot occur.')
    elif inputSheet.title == 'TU130' and inputSheet.max_column != 6:
        dataErrors.append('An incorrect number of columns are present in the spreadsheet. Further validation cannot occur.')
    else:
        for row in range(4,inputSheet.max_row + 1): #begin validating data, row by row   
            for column in range(1, inputSheet.max_column + 1):
                currentCellValue = inputSheet.cell(row = row, column = column).value
                if currentCellValue == None: #Blank cell check
                    dataErrors.append('A blank cell is present. Row ' + str(row) + '.')
                    continue
                if inputSheet.title == 'T025X':
                    populateSubsets(conn, sqlT024X1, subsetT024X1, field24X1)
                    populateSubsets(conn, sqlTAD3F, subsetTAD3F, fieldD3F)
                    populateSubsets(conn, sqlT0C8X1, subsetT0C8X1, fieldT0C8X1)
                    populateSubsets(conn, sqlT0C8X2, subsetT0C8X2, fieldT0C8X2)
                    
                    if column == 1: #COMPANY_CODE check
                        dataErrors.extend(validateCompanyCode(inputSheet, row, column))
                    if column == 2: #PRODUCT_PREFIX check
                        dataErrors.extend(validateProductPrefix(inputSheet, row, column))
                    if column == 3: #TABLE_SUBSET check
                        dataErrors.extend(validateTableSubset(currentCellValue, row, subsetT024X1, subsetTAD3F, subsetT0C8X1, subsetT0C8X2))
                    if column == 4: #ISSUE_STATE check
                        dataErrors.extend(validateIssueState(inputSheet, row, column))
                    if column == 5: #RCPT_PERD_STRT_DT check
                        try:
                             startDate = str(inputSheet.cell(row = row, column = column).value)
                             datetime.datetime.strptime(startDate, '%Y-%m-%d')
                        except:
                             dataErrors.append('The RCPT_PERD_STRT_DT record is invalid. Row ' + str(row) + '.')
                    if column == 6: #INT_RT_EFF_DT check
                        try:
                            effectiveDate = str(inputSheet.cell(row = row, column = column).value)
                            datetime.datetime.strptime(effectiveDate, '%Y-%m-%d')
                        except:
                            dataErrors.append('The INT_RT_EFF_DT record is invalid. Row ' + str(row) + '.')
                    if column == 7: #SETTL_DATE_IND check
                        if len(str(inputSheet.cell(row = row, column = column).value)) > 1:
                            dataErrors.append('The SETTL_DATE_IND record is too long. Row ' + str(row) + '.')
                    if column == 8: #INTEREST_RATE check
                        if re.search(r'^\d{1,2}(\.\d{0,3})?$', str(inputSheet.cell(row = row, column = column).value)) == None:
                            dataErrors.append('An invalidly formatted INTEREST_RATE record is present. Row ' + str(row) + '.')
                        elif float(inputSheet.cell(row = row, column = column).value) < 0 or float(inputSheet.cell(row = row, column = column).value) > 99.999:
                            dataErrors.append('An INTEREST_RATE is outside of the acceptable threshold. Row ' + str(row) + '.')
                    if column == 9: #ACTION check
                        if inputSheet.cell(row = row, column = column).value not in ['U', 'A']: #Valid action check
                            dataErrors.append('The ACTION record is invalid. Row ' + str(row) + '.')

                        if inputSheet.cell(row = row, column = column).value == 'A': #Duplicate add row check
                            for i in range(0,8):
                                rowData.append(inputSheet.cell(row = row, column = i + 1).value)
                            if rowData in allAddRows:
                                dataErrors.append('A duplicate row is present. Row ' + str(row) + '.')
                            allAddRows.append(rowData)
                            rowData = []

                elif validateTable(inputSheet) == 'T026X': #Run validation on the inputSheet, based on table selection.
                    populateSubsets(conn, sqlT024X2, subsetT024X2, field24X2)
                    populateSubsets(conn, sqlT0C8X3, subsetT0C8X3, fieldT0C8X3)
                    populateSubsets(conn, sqlT0C8X4, subsetT0C8X4, fieldT0C8X4)
                    populateSubsets(conn, sqlT0C8X5, subsetT0C8X5, fieldT0C8X5)
                    
                    if column == 1: #COMPANY_CODE check
                        dataErrors.extend(validateCompanyCode(inputSheet, row, column))
                    if column == 2: #PRODUCT_PREFIX check
                        dataErrors.extend(validateProductPrefix(inputSheet, row, column))
                    if column == 3: #TABLE_SUBSET check
                        if len(str(inputSheet.cell(row = row, column = column).value)) > 16:
                            dataErrors.append('The TABLE_SUBSET record is too long. Row ' + str(row) + '.')
                    if column == 4: #ISSUE_STATE check
                        dataErrors.extend(validateIssueState(inputSheet, row, column))
                    if column == 5: #EFFECTIVE_DATE check
                        try:
                            effectiveDate = str(inputSheet.cell(row = row, column = column).value)
                            datetime.datetime.strptime(effectiveDate, '%Y-%m-%d')
                        except:
                            dataErrors.append('The EFFECTIVE_DATE record is invalid. Row ' + str(row) + '.')
                    if column == 6: #MAXIMUM_DURATION check
                        if len(str(inputSheet.cell(row = row, column = column).value)) > 3:
                            dataErrors.append('The MAXIMUM_DURATION record is too long. Row ' + str(row) + '.')
                        else:
                            try:
                                inputSheet.cell(row = row, column = column).value + 1
                            except:
                                dataErrors.append('The MAXIMUM_DURATION record is incorrectly formatted. Row ' + str(row) + '.')
                    if column == 7: #MX_CAL_YY check
                        if len(str(inputSheet.cell(row = row, column = column).value)) > 5:
                            dataErrors.append('The MX_CAL_YY record is too long. Row ' + str(row) + '.')
                        else:
                            try:
                                inputSheet.cell(row = row, column = column).value + 1
                            except:
                                dataErrors.append('The MX_CAL_YY record is incorrectly formatted. Row ' + str(row) + '.')
                    if column == 8: #GUAR_INT_RT check
                        if re.search(r'^\d{1,2}\.\d{0,3}$', str(inputSheet.cell(row = row, column = column).value)) == None:
                            dataErrors.append('An invalidly formatted GUAR_INT_RT record is present. Row ' + str(row) + '.')
                        elif float(inputSheet.cell(row = row, column = column).value) < 0 or float(inputSheet.cell(row = row, column = column).value) > 99.999:
                            dataErrors.append('An GUAR_INT_RT is outside of the acceptable threshold. Row ' + str(row) + '.')
                    if column == 9: #ACTION check
                        if inputSheet.cell(row = row, column = column).value not in ['U', 'A']: #Valid action check
                            dataErrors.append('The ACTION record is invalid. Row ' + str(row) + '.')

                        if inputSheet.cell(row = row, column = column).value == 'A': #Duplicate add row check
                            for i in range(0,8):
                                rowData.append(inputSheet.cell(row = row, column = i + 1).value)
                            if rowData in allAddRows:
                                dataErrors.append('A duplicate row is present. Row ' + str(row) + '.')
                            allAddRows.append(rowData)
                            rowData = []

                elif validateTable(inputSheet) == 'TU130':
                    if column == 1: #COMPANY_CODE check
                        dataErrors.extend(validateCompanyCode(inputSheet, row, column))
                    if column == 2: #INDEX_TYPE field check
                        if len(str(inputSheet.cell(row = row, column = column).value)) > 3:
                            dataErrors.append('The INDEX_TYPE record is too long. Row ' + str(row) + '.')
                    if column == 3: #EFFECTIVE_DATE check:
                        try:
                            effectiveDate = str(inputSheet.cell(row = row, column = column).value)
                            datetime.datetime.strptime(effectiveDate, '%Y-%m-%d')
                        except:
                            dataErrors.append('The EFFECTIVE_DATE record is invalid. Row ' + str(row) + '.')
                    if column == 4: #GUAR_PERIOD check
                        if re.search(r'^\d{1,4}$', str(inputSheet.cell(row = row, column = column).value)) == None:
                            dataErrors.append('An invalidly formatted GUAR_PERIOD record is present. Row ' + str(row) + '.')
                    if column == 5: #INDEX_RATE checks
                        if re.search(r'^\.\d{1,5}$', str(inputSheet.cell(row = row, column = column).value)) == None:
                            dataErrors.append('An invalidly formatted INDEX_RATE record is present. Row ' + str(row) + '.')
                        elif float(inputSheet.cell(row = row, column = 5).value) < 0 or float(inputSheet.cell(row = row, column = column).value) > 0.25:
                            dataErrors.append('An INDEX_RATE is outside of the acceptable threshold. Row ' + str(row) + '.')
                    if column == 6: #ACTION check
                        if inputSheet.cell(row = row, column = column).value not in ['U', 'A']: #valid action check
                            dataErrors.append('The ACTION record is invalid. Row ' + str(row) + '.')
                        if inputSheet.cell(row = row, column = column).value == 'A': #Duplicate add row check
                            for i in range(0,5):
                                rowData.append(inputSheet.cell(row = row, column = i + 1).value)
                            if rowData in allAddRows:
                                dataErrors.append('A duplicate row is present. Row ' + str(row) + '.')
                            allAddRows.append(rowData)
                            rowData = []
                            
    return dataErrors

def validateSpreadsheet(inputXLSX):
    errorMessages = []
    errorMessageString = ''

    print('Opening workbook...') #Take an input CSV.
    wb = openpyxl.load_workbook(inputXLSX) #Change this to test other tables
    sheet = wb.active
    print('Workbook opened.')

    if validateTable(sheet) == None:
        return('An incomplete control record is present or the spreadsheet is blank.')
    elif validateTable(sheet) == False:
        return('The worksheet name must be T025X, T026X, or TU130. \n')
    else:
        print('Beginning validation for ' + sheet.title + '...')
        errorMessages.extend(validateControlRecord(sheet)) #Gather control errors
        errorMessages.extend(validateData(sheet)) #Gather data errors
        errorMessageString = printErrorMessages(errorMessages)
        #errorMessageString = ''.join(map(str, errorMessages))
        return(errorMessageString) #Print error messages
